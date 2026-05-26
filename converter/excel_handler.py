import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def flatten_contact(contact: dict) -> dict:
    """Flattens a structured contact dictionary into a flat key-value dict for Excel columns."""
    flat = {
        'Display Name': contact.get('display_name', '') or '',
        'First Name': contact.get('first_name', '') or '',
        'Middle Name': contact.get('middle_name', '') or '',
        'Last Name': contact.get('last_name', '') or '',
        'Prefix': contact.get('prefix', '') or '',
        'Suffix': contact.get('suffix', '') or '',
        'Organization': contact.get('org', '') or '',
        'Department': contact.get('department', '') or '',
        'Job Title': contact.get('title', '') or '',
        'Birthday': contact.get('bday', '') or '',
        'Notes': contact.get('note', '') or ''
    }
    
    # Process phones
    phone_counts = {}
    for phone in contact.get('phones', []):
        val = phone.get('value', '').strip()
        if not val:
            continue
        types = phone.get('types', [])
        label = ' '.join(types).title() if types else ''
        key_prefix = f"Phone - {label}".strip() if label else "Phone"
        
        phone_counts[key_prefix] = phone_counts.get(key_prefix, 0) + 1
        col_name = f"{key_prefix} {phone_counts[key_prefix]}"
        flat[col_name] = val
        
    # Process emails
    email_counts = {}
    for email in contact.get('emails', []):
        val = email.get('value', '').strip()
        if not val:
            continue
        types = email.get('types', [])
        label = ' '.join(types).title() if types else ''
        key_prefix = f"Email - {label}".strip() if label else "Email"
        
        email_counts[key_prefix] = email_counts.get(key_prefix, 0) + 1
        col_name = f"{key_prefix} {email_counts[key_prefix]}"
        flat[col_name] = val
        
    # Process URLs
    url_counts = {}
    for url in contact.get('urls', []):
        val = url.get('value', '').strip()
        if not val:
            continue
        types = url.get('types', [])
        label = ' '.join(types).title() if types else ''
        key_prefix = f"URL - {label}".strip() if label else "URL"
        
        url_counts[key_prefix] = url_counts.get(key_prefix, 0) + 1
        col_name = f"{key_prefix} {url_counts[key_prefix]}"
        flat[col_name] = val
        
    # Process addresses
    addr_counts = {}
    for addr in contact.get('addresses', []):
        types = addr.get('types', [])
        label = ' '.join(types).title() if types else ''
        key_prefix = f"Address - {label}".strip() if label else "Address"
        
        addr_counts[key_prefix] = addr_counts.get(key_prefix, 0) + 1
        idx = addr_counts[key_prefix]
        
        # Add street, city, region, zip, country
        for field in ['street', 'city', 'region', 'zip', 'country']:
            val = addr.get(field, '').strip()
            col_name = f"{key_prefix} {idx} - {field.title()}"
            flat[col_name] = val or ''
            
    return flat

def flatten_contacts(contacts: list[dict]) -> tuple[list[dict], list[str]]:
    """Flattens a list of contacts and returns the flat dict list and an ordered list of headers."""
    flat_list = []
    all_keys = set()
    for c in contacts:
        flat_c = flatten_contact(c)
        flat_list.append(flat_c)
        all_keys.update(flat_c.keys())
        
    # Base columns in their logical ordering
    base_cols = [
        'Display Name', 'First Name', 'Middle Name', 'Last Name', 'Prefix', 'Suffix',
        'Organization', 'Department', 'Job Title', 'Birthday', 'Notes'
    ]
    
    # Filter keys that have at least one value across all contacts
    active_keys = set()
    for flat_c in flat_list:
        for k, v in flat_c.items():
            if v and str(v).strip():
                active_keys.add(k)
                
    ordered_cols = []
    # Add base columns if they have active values
    for col in base_cols:
        if col in active_keys:
            ordered_cols.append(col)
            
    # Add other dynamic columns (Phones, Emails, URLs, Addresses) sorted alphabetically
    other_keys = sorted(list(all_keys - set(base_cols)))
    for col in other_keys:
        if col in active_keys:
            ordered_cols.append(col)
            
    # Fallback to base columns if everything is empty
    if not ordered_cols:
        ordered_cols = base_cols
        
    return flat_list, ordered_cols

def contacts_to_excel(contacts: list[dict], filepath: str):
    """Writes a list of contact dictionaries to an Excel file with styled header and auto column widths."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    
    # Flatten and get headers
    flat_list, columns = flatten_contacts(contacts)
    
    # Header Styling
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write Header
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    # Body Styling
    body_font = Font(name="Segoe UI", size=10)
    body_align = Alignment(vertical="center")
    
    # Write Rows
    for row_idx, flat_c in enumerate(flat_list, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = flat_c.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = body_align
            
    # Auto Column Dimensions
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Safe margins for autowidth
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 45), 12)
        
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    
    wb.save(filepath)

def auto_map_columns(headers: list[str]) -> dict:
    """Inspects Excel headers and automatically maps them to contact field identifiers."""
    mapping = {}
    for h in headers:
        h_clean = h.strip().lower()
        if not h_clean:
            continue
            
        # Exact/Alias checks
        if 'display' in h_clean or 'full' in h_clean or h_clean == 'name':
            mapping[h] = 'display_name'
        elif 'first' in h_clean or 'given' in h_clean:
            mapping[h] = 'first_name'
        elif 'last' in h_clean or 'family' in h_clean or 'surname' in h_clean:
            mapping[h] = 'last_name'
        elif 'middle' in h_clean:
            mapping[h] = 'middle_name'
        elif 'prefix' in h_clean or 'title' in h_clean and ('mr' in h_clean or 'honorific' in h_clean):
            mapping[h] = 'prefix'
        elif 'suffix' in h_clean:
            mapping[h] = 'suffix'
        elif 'org' in h_clean or 'company' in h_clean or 'corporation' in h_clean or 'business' in h_clean:
            mapping[h] = 'org'
        elif 'dept' in h_clean or 'department' in h_clean:
            mapping[h] = 'department'
        elif 'job' in h_clean or 'title' in h_clean or 'position' in h_clean or 'role' in h_clean:
            mapping[h] = 'title'
        elif 'bday' in h_clean or 'birth' in h_clean:
            mapping[h] = 'bday'
        elif 'note' in h_clean or 'memo' in h_clean or 'comment' in h_clean or 'desc' in h_clean:
            mapping[h] = 'note'
            
        # Phones
        elif 'phone' in h_clean or 'tel' in h_clean or 'mobile' in h_clean or 'cell' in h_clean or 'contact' in h_clean:
            if 'cell' in h_clean or 'mobile' in h_clean or 'mobi' in h_clean:
                mapping[h] = 'phone_cell'
            elif 'work' in h_clean:
                mapping[h] = 'phone_work'
            elif 'home' in h_clean:
                mapping[h] = 'phone_home'
            elif 'fax' in h_clean:
                mapping[h] = 'phone_fax'
            elif 'pager' in h_clean:
                mapping[h] = 'phone_pager'
            else:
                mapping[h] = 'phone_other'
                
        # Emails
        elif 'email' in h_clean or 'mail' in h_clean:
            if 'work' in h_clean:
                mapping[h] = 'email_work'
            elif 'home' in h_clean:
                mapping[h] = 'email_home'
            else:
                mapping[h] = 'email_other'
                
        # URLs
        elif 'url' in h_clean or 'web' in h_clean or 'site' in h_clean or 'homepage' in h_clean:
            if 'work' in h_clean:
                mapping[h] = 'url_work'
            elif 'home' in h_clean:
                mapping[h] = 'url_home'
            else:
                mapping[h] = 'url_other'
                
        # Addresses
        elif 'address' in h_clean or 'addr' in h_clean or 'street' in h_clean or 'city' in h_clean or 'zip' in h_clean or 'country' in h_clean or 'region' in h_clean or 'state' in h_clean:
            # Determine address type
            a_type = 'other'
            if 'work' in h_clean:
                a_type = 'work'
            elif 'home' in h_clean:
                a_type = 'home'
                
            # Determine field
            if 'street' in h_clean or 'addr' in h_clean or 'line' in h_clean:
                mapping[h] = f'address_{a_type}_street'
            elif 'city' in h_clean or 'town' in h_clean:
                mapping[h] = f'address_{a_type}_city'
            elif 'state' in h_clean or 'region' in h_clean or 'province' in h_clean:
                mapping[h] = f'address_{a_type}_region'
            elif 'zip' in h_clean or 'postal' in h_clean or 'post' in h_clean:
                mapping[h] = f'address_{a_type}_zip'
            elif 'country' in h_clean or 'nation' in h_clean:
                mapping[h] = f'address_{a_type}_country'
            else:
                mapping[h] = f'address_{a_type}_street' # Fallback
        else:
            # Default fallback mapping based on common fields, or leave unmapped
            pass
            
    return mapping

def excel_to_contacts(filepath: str, mapping: dict) -> list[dict]:
    """Reads contacts from Excel file using the provided column mapping."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if not wb.sheetnames:
        return []
    ws = wb.active
    
    # Extract headers
    headers = []
    # Read first row
    for row in ws.iter_rows(max_row=1):
        for cell in row:
            headers.append(str(cell.value or '').strip())
            
    contacts = []
    
    # Read subsequent rows
    for row in ws.iter_rows(min_row=2):
        row_vals = [cell.value for cell in row]
        if not any(val is not None for val in row_vals):
            continue # Skip empty rows
            
        contact = {
            'first_name': '',
            'middle_name': '',
            'last_name': '',
            'prefix': '',
            'suffix': '',
            'display_name': '',
            'org': '',
            'department': '',
            'title': '',
            'bday': '',
            'note': '',
            'phones': [],
            'emails': [],
            'addresses': [],
            'urls': [],
            'custom': []
        }
        
        phones_by_type = {}
        emails_by_type = {}
        urls_by_type = {}
        addr_by_type = {}
        
        for idx, val in enumerate(row_vals):
            if idx >= len(headers):
                break
            col_name = headers[idx]
            target = mapping.get(col_name)
            if not target or val is None:
                continue
                
            val_str = str(val).strip()
            if not val_str:
                continue
                
            if target == 'first_name':
                contact['first_name'] = val_str
            elif target == 'last_name':
                contact['last_name'] = val_str
            elif target == 'middle_name':
                contact['middle_name'] = val_str
            elif target == 'prefix':
                contact['prefix'] = val_str
            elif target == 'suffix':
                contact['suffix'] = val_str
            elif target == 'display_name':
                contact['display_name'] = val_str
            elif target == 'org':
                contact['org'] = val_str
            elif target == 'department':
                contact['department'] = val_str
            elif target == 'title':
                contact['title'] = val_str
            elif target == 'bday':
                contact['bday'] = val_str
            elif target == 'note':
                contact['note'] = val_str
            elif target.startswith('phone_'):
                p_type = target.split('_', 1)[1].upper()
                if p_type == 'OTHER':
                    p_type = ''
                phones_by_type.setdefault(p_type, []).append(val_str)
            elif target.startswith('email_'):
                e_type = target.split('_', 1)[1].upper()
                if e_type == 'OTHER':
                    e_type = ''
                emails_by_type.setdefault(e_type, []).append(val_str)
            elif target.startswith('url_'):
                u_type = target.split('_', 1)[1].upper()
                if u_type == 'OTHER':
                    u_type = ''
                urls_by_type.setdefault(u_type, []).append(val_str)
            elif target.startswith('address_'):
                parts = target.split('_')
                if len(parts) >= 3:
                    a_type = parts[1].upper()
                    a_field = parts[2] # street, city, region, zip, country
                    if a_type == 'OTHER':
                        a_type = ''
                    addr_by_type.setdefault(a_type, {})[a_field] = val_str
                    
        # Compile Phones
        for p_type, values in phones_by_type.items():
            for v in values:
                # Check if multiple values are stored in single cell comma/semicolon-separated
                for sub_v in re.split(r'[;,]', v):
                    sub_v = sub_v.strip()
                    if sub_v:
                        contact['phones'].append({
                            'value': sub_v,
                            'types': [p_type] if p_type else []
                        })
                        
        # Compile Emails
        for e_type, values in emails_by_type.items():
            for v in values:
                for sub_v in re.split(r'[;,]', v):
                    sub_v = sub_v.strip()
                    if sub_v:
                        contact['emails'].append({
                            'value': sub_v,
                            'types': [e_type] if e_type else []
                        })
                        
        # Compile URLs
        for u_type, values in urls_by_type.items():
            for v in values:
                for sub_v in re.split(r'[;,]', v):
                    sub_v = sub_v.strip()
                    if sub_v:
                        contact['urls'].append({
                            'value': sub_v,
                            'types': [u_type] if u_type else []
                        })
                        
        # Compile Addresses
        for a_type, fields in addr_by_type.items():
            if any(fields.values()):
                contact['addresses'].append({
                    'street': fields.get('street', ''),
                    'city': fields.get('city', ''),
                    'region': fields.get('region', ''),
                    'zip': fields.get('zip', ''),
                    'country': fields.get('country', ''),
                    'types': [a_type] if a_type else []
                })
                
        # Generate Display Name if empty
        if not contact['display_name'] and (contact['first_name'] or contact['last_name']):
            parts = [contact['prefix'], contact['first_name'], contact['middle_name'], contact['last_name'], contact['suffix']]
            contact['display_name'] = ' '.join([p for p in parts if p])
            
        contacts.append(contact)
        
    return contacts
