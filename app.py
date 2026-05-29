import io
import json
import logging
import os
import socket
import webbrowser
from threading import Timer
from flask import Flask, request, jsonify, send_file, render_template
from converter import parse_vcf, write_vcf, contacts_to_excel, excel_to_contacts, auto_map_columns
import openpyxl

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Add file handler to write server log file
try:
    file_handler = logging.FileHandler('server.log', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
    logging.getLogger().addHandler(file_handler)
    logger.info("Logging to server.log initialized.")
except Exception as e:
    print(f"Failed to initialize file logging: {str(e)}")

app = Flask(__name__, template_folder='templates', static_folder='static')

# Limit upload file sizes to 16MB
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

@app.route('/')
def index():
    """Serves the single-page GUI."""
    return render_template('index.html')

@app.route('/api/vcf-to-json', methods=['POST'])
def vcf_to_json():
    """Parses an uploaded VCF file and returns JSON list of contacts."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
        
    file = request.files['file']
    if not file.filename.endswith('.vcf'):
        return jsonify({'error': 'Invalid file format. Must be a .vcf file'}), 400
        
    try:
        content = file.read().decode('utf-8', errors='ignore')
        contacts = parse_vcf(content)
        return jsonify({
            'success': True,
            'contacts': contacts,
            'count': len(contacts)
        })
    except Exception as e:
        logger.error(f"Error parsing VCF: {str(e)}", exc_info=True)
        return jsonify({'error': f"Failed to parse VCF file: {str(e)}"}), 500

@app.route('/api/json-to-excel', methods=['POST'])
def json_to_excel():
    """Converts a JSON list of contacts into an Excel file and returns it for download."""
    try:
        data = request.get_json()
        if not data or 'contacts' not in data:
            return jsonify({'error': 'No contacts data provided'}), 400
            
        contacts = data['contacts']
        if not contacts:
            return jsonify({'error': 'Contacts list is empty'}), 400
            
        # Create in-memory file
        excel_buffer = io.BytesIO()
        contacts_to_excel(contacts, excel_buffer)
        excel_buffer.seek(0)
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='contacts_converted.xlsx'
        )
    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}", exc_info=True)
        return jsonify({'error': f"Failed to generate Excel: {str(e)}"}), 500

@app.route('/api/excel-to-json', methods=['POST'])
def excel_to_json():
    """Reads headers and preview rows from Excel file, returns them along with auto-mapped fields."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
        
    file = request.files['file']
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'error': 'Invalid file format. Must be an Excel file (.xlsx)'}), 400
        
    try:
        # Load in memory
        in_memory_file = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(in_memory_file, read_only=True, data_only=True)
        if not wb.sheetnames:
            return jsonify({'error': 'Excel file contains no sheets'}), 400
            
        ws = wb.active
        
        # Read headers
        headers = []
        for row in ws.iter_rows(max_row=1):
            for cell in row:
                headers.append(str(cell.value or '').strip())
                
        # Filter empty headers
        headers = [h for h in headers if h]
        if not headers:
            return jsonify({'error': 'Excel file has no headers in the first row'}), 400
            
        # Read a preview of first 5 rows
        preview_rows = []
        for row in ws.iter_rows(min_row=2, max_row=6):
            row_vals = [cell.value for cell in row]
            # Strip to match length of headers
            row_vals = row_vals[:len(headers)]
            # Fill with empty string if shorter
            row_vals += [''] * (len(headers) - len(row_vals))
            # If row is empty, skip
            if not any(val is not None and str(val).strip() for val in row_vals):
                continue
            preview_rows.append({headers[i]: str(row_vals[i] or '') for i in range(len(headers))})
            
        # Get auto mapping
        mapping = auto_map_columns(headers)
        
        return jsonify({
            'success': True,
            'headers': headers,
            'preview': preview_rows,
            'auto_mapping': mapping
        })
    except Exception as e:
        logger.error(f"Error reading Excel: {str(e)}", exc_info=True)
        return jsonify({'error': f"Failed to read Excel file: {str(e)}"}), 500

@app.route('/api/excel-to-vcf', methods=['POST'])
def excel_to_vcf_convert():
    """Converts Excel file to VCF using the supplied column mapping."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    if 'mapping' not in request.form:
        return jsonify({'error': 'No column mapping provided'}), 400
        
    file = request.files['file']
    try:
        mapping = json.loads(request.form['mapping'])
    except Exception:
        return jsonify({'error': 'Invalid mapping format'}), 400
        
    try:
        in_memory_file = io.BytesIO(file.read())
        # Parse contacts from Excel
        contacts = excel_to_contacts(in_memory_file, mapping)
        
        if not contacts:
            return jsonify({'error': 'No contacts parsed from Excel file'}), 400
            
        # Generate VCF string
        vcf_str = write_vcf(contacts)
        
        vcf_buffer = io.BytesIO(vcf_str.encode('utf-8'))
        vcf_buffer.seek(0)
        
        return send_file(
            vcf_buffer,
            mimetype='text/vcard',
            as_attachment=True,
            download_name='contacts_converted.vcf'
        )
    except Exception as e:
        logger.error(f"Error converting Excel to VCF: {str(e)}", exc_info=True)
        return jsonify({'error': f"Failed to convert: {str(e)}"}), 500

def find_free_port(start_port=5000):
    """Finds an available TCP port starting from start_port."""
    port = start_port
    while port < 6000:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(('127.0.0.1', port))
                return port
            except OSError:
                port += 1
    return start_port

def open_browser(port):
    """Opens the local web application in the default browser."""
    url = f"http://127.0.0.1:{port}"
    try:
        if os.name == 'nt':
            os.system(f'start "" "{url}"')
        else:
            webbrowser.open(url)
    except Exception as e:
        logger.error(f"Failed to open browser automatically: {str(e)}")

if __name__ == '__main__':
    port = find_free_port(5000)
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    
    # Start browser timer in parent process only (if in debug/reload mode)
    if not debug_mode or not os.environ.get("WERKZEUG_RUN_MAIN"):
        Timer(1.0, open_browser, [port]).start()
        
    logger.info(f"Starting server on http://127.0.0.1:{port}")
    app.run(host='127.0.0.1', port=port, debug=debug_mode)
